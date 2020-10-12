# -*-coding:utf-8-*-

import os
from openpyxl import load_workbook


class ExcelUtil(object):
    def __init__(self, excel_path=None, sheet_name=None):
        """获取 excel 工作表"""

        if excel_path is None:
            current_path = os.path.abspath(os.path.dirname(__file__))
            self.excel_path = current_path + '/../case/casedata.xlsx'

        else:
            self.excel_path = excel_path

        if sheet_name is None:
            self.sheet_name = "Sheet"
        else:
            self.sheet_name = sheet_name

        # 打开工作表
        self.data = load_workbook(self.excel_path)
        self.sheet = self.data[self.sheet_name]

    def get_data(self):
        """
        获取文件数据
        每一行数据一个list，所有的数据一个大list
        """

        rows = self.sheet.rows
        row_num = self.sheet.max_row
        col_num = self.sheet.max_column

        if row_num <= 1:
            print("总行数小于1,没有数据")
        else:
            case_all = []
            for row in rows:
                case = []
                for i in range(col_num):
                    case.append(row[i].value)
                case_all.append(case)
            return case_all

    def get_case(self):
        """
        获取数据
        得到有用的数据，并且使数据以对象、操作、参数的顺序返回
        :return: [[element, action, par],[element, action, par]...]
        """
        data = self.get_data()

        # 得到所需要数据的索引，然后根据索引获取相应顺序的数据
        element_index = data[0].index("对象")
        action_index = data[0].index("操作")
        parameter_index = data[0].index("参数")

        all_case = []
        # 去除header行，和其他无用的数据
        for i in range(1, len(data)):
            case = []
            case.append(data[i][element_index])
            case.append(data[i][action_index])
            case.append(data[i][parameter_index])
            all_case.append(case)
        return all_case


if __name__ == '__main__':
    sheet = '技术顾问预约'
    file = ExcelUtil(sheet_name=sheet)
    print(file.get_case())
